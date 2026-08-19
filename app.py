import io
import os
import re
from datetime import date

import pandas as pd
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import xlrd
except ImportError:
    xlrd = None

st.set_page_config(page_title="행사제안서 자동생성기", layout="wide")

# ---------------------------------------------------------------
# 기준 파일을 서버(로컬 실행 폴더)에 저장해 두고 재사용하기 위한 경로
# ---------------------------------------------------------------
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
COST_DIR = os.path.join(DATA_DIR, "cost")
PRICE_DIR = os.path.join(DATA_DIR, "price")
CHANNEL_DIR = os.path.join(DATA_DIR, "channel")
LOG_PATH = os.path.join(DATA_DIR, "download_log.csv")
os.makedirs(COST_DIR, exist_ok=True)
os.makedirs(PRICE_DIR, exist_ok=True)
os.makedirs(CHANNEL_DIR, exist_ok=True)

# ---------------------------------------------------------------
# 고정 컬럼 위치 (0-indexed, A=0)
# ---------------------------------------------------------------
COL = {"E": 4, "F": 5, "G": 6, "H": 7, "I": 8, "J": 9, "M": 12, "N": 13}

EVENT_OPTIONS = [
    ("최저가", "최저가"),
    ("상시가", "상시할인가"),
    ("day7", "7일 행사가(폐쇄몰)"),
    ("day3", "3일 행사가"),
    ("공동구매가", "공동구매가"),
    ("기타", "기타(직접입력)"),
]
EVENT_LABEL = dict(EVENT_OPTIONS)
EVENT_KEYS = [k for k, _ in EVENT_OPTIONS]

# 단품 판단에서 "묶음/세트"로 간주할 보조 키워드 (수량 숫자 패턴이 없을 때만 사용)
BUNDLE_HINT_WORDS = ["세트", "SET", "묶음", "구성"]


# ---------------------------------------------------------------
# 유틸
# ---------------------------------------------------------------
def get(row, idx):
    if idx is None:
        return ""
    if idx < len(row):
        v = row[idx]
        return "" if v is None else v
    return ""


def to_num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("₩", "").strip()
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def normalize_code(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s == "":
        return ""
    try:
        f = float(s)
        if re.match(r"^-?\d+(\.0+)?$", s):
            return str(int(f))
    except ValueError:
        pass
    return s


def extract_qty(row):
    """행 전체 텍스트에서 수량(개입) 힌트를 찾아 단품 여부를 판단.
    '1개'면 1, '2개입'/'3개' 등이면 그 숫자, 아무 힌트가 없으면 단품(1)으로 간주."""
    text = " ".join(str(c) for c in row if c not in (None, ""))
    nums = [int(m) for m in re.findall(r"(\d+)\s*개", text)]
    if nums:
        return min(nums)
    if any(w in text for w in BUNDLE_HINT_WORDS):
        return 99  # 세트/묶음 힌트만 있고 숫자가 없으면 단품이 아닌 것으로 취급
    return 1


def read_rows(file_bytes, ext):
    """엑셀 파일을 항상 A열=0 기준 절대 위치의 2차원 리스트로 읽는다."""
    if ext == "xlsx":
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        return rows
    elif ext == "xls":
        if xlrd is None:
            raise RuntimeError("xlrd 라이브러리가 설치되어 있지 않습니다 (requirements.txt 확인).")
        wb = xlrd.open_workbook(file_contents=file_bytes)
        sheet = wb.sheet_by_index(0)
        rows = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
        return rows
    else:
        raise RuntimeError("지원하지 않는 파일 형식입니다 (.xls 또는 .xlsx만 가능).")


def find_header_row(rows, keywords, min_matches=2):
    best_row, best_map = -1, {}
    for r in range(min(len(rows), 20)):
        row = rows[r]
        col_map = {}
        for ci, cell in enumerate(row):
            v = str(cell).strip() if cell is not None else ""
            for k in keywords:
                if k in v and k not in col_map:
                    col_map[k] = ci
        if len(col_map) >= min_matches:
            return r, col_map
        if len(col_map) > len(best_map):
            best_row, best_map = r, col_map
    return best_row, best_map


def parse_cost_sheet(rows):
    """.xls 원가 파일: 품번, 품명, 원가는 헤더 텍스트로 탐색.
    같은 품번이 여러 줄(묶음 수량별)로 존재하면 단품(1개) 기준 행을 우선 채택."""
    header_row, col_map = find_header_row(rows, ["품번", "품명", "원가"], 2)
    result = {}
    if header_row == -1:
        return result
    c_name = col_map.get("품명")
    c_code = col_map.get("품번")
    c_cost = col_map.get("원가")
    for r in range(header_row + 1, len(rows)):
        row = rows[r]
        if not row:
            continue
        code = normalize_code(get(row, c_code)) if c_code is not None else ""
        if not code:
            continue
        name = str(get(row, c_name)).strip() if c_name is not None else ""
        cost_val = to_num(get(row, c_cost)) if c_cost is not None else None
        qty = extract_qty(row)
        if code not in result or qty < result[code]["_qty"]:
            result[code] = {"품명": name, "원가": cost_val, "_qty": qty}
    return result


def parse_price_sheet(rows):
    """.xlsx 가격표: 상품코드 컬럼은 헤더 탐색, E~J/M/N은 고정 위치.
    같은 상품코드가 여러 줄(묶음 수량별)로 존재하면 단품(1개) 기준 행을 우선 채택."""
    header_row, col_map = find_header_row(rows, ["상품코드", "품번"], 1)
    if header_row == -1:
        header_row = 0
    c_code = col_map.get("상품코드", col_map.get("품번", 0))
    result = {}
    for r in range(header_row + 1, len(rows)):
        row = rows[r]
        if not row:
            continue
        code = normalize_code(get(row, c_code))
        if not code:
            continue
        qty = extract_qty(row)
        if code in result and qty >= result[code]["_qty"]:
            continue  # 이미 더 단품에 가까운 행을 채택한 상태면 건너뜀
        rocket_raw = str(get(row, COL["M"])).strip()
        arrival_raw = str(get(row, COL["N"])).strip()
        result[code] = {
            "정상가": to_num(get(row, COL["E"])),
            "최저가": to_num(get(row, COL["F"])),
            "상시가": to_num(get(row, COL["G"])),
            "day7": to_num(get(row, COL["H"])),
            "day3": to_num(get(row, COL["I"])),
            "공동구매가": to_num(get(row, COL["J"])),
            # 로켓배송: 원본 값(운영/위수탁 등)을 그대로 표기, 공란이면 '-'
            "rocket": rocket_raw if rocket_raw not in ("", "None") else "-",
            "arrival": "도착보장" if arrival_raw not in ("", "None") else "-",
            "_qty": qty,
        }
    return result


def parse_channel_sheet(rows):
    """판매처별 매핑 파일: 오클릭 품번 <-> 어드민 상품코드.
    헤더 텍스트에서 '오클릭'이 포함된 컬럼을 오클릭 품번, '어드민'+'상품코드'가 포함된 컬럼을
    어드민 상품코드로 인식한다. 못 찾으면 A열=어드민 상품코드, B열=오클릭 품번으로 가정."""
    header_row, _ = find_header_row(rows, ["오클릭", "어드민", "상품코드", "품번"], 2)
    if header_row == -1:
        header_row = 0
    hrow = rows[header_row] if header_row < len(rows) else []
    c_admin, c_oclick = None, None
    for ci, cell in enumerate(hrow):
        v = str(cell).strip() if cell is not None else ""
        if "오클릭" in v and c_oclick is None:
            c_oclick = ci
        if "어드민" in v and "상품코드" in v and c_admin is None:
            c_admin = ci
    if c_admin is None:
        c_admin = 0
    if c_oclick is None:
        c_oclick = 1
    result = {}
    for r in range(header_row + 1, len(rows)):
        row = rows[r]
        if not row:
            continue
        oclick_code = normalize_code(get(row, c_oclick))
        if not oclick_code:
            continue
        admin_code = str(get(row, c_admin)).strip() if get(row, c_admin) != "" else ""
        if admin_code:
            result[oclick_code] = admin_code
    return result


def calc_row(r, cost_raw):
    event_price = r["기타금액"] if r["적용타입"] == "기타" else r.get(r["적용타입"])
    price_missing = event_price is None  # 선택한 항목의 금액이 비어있음
    discount_rate = (1 - event_price / r["정상가"]) if r.get("정상가") and event_price is not None else None
    discount_amt = (r["정상가"] - event_price) if r.get("정상가") is not None and event_price is not None else None
    cost_vat_in = cost_raw
    cost_vat_ex = round(cost_raw / 1.1) if cost_raw is not None else None
    fee_dec = (r.get("수수료") or 0) / 100
    supply_vat_in = round(event_price * (1 - fee_dec)) if event_price is not None else None
    supply_vat_ex = round(supply_vat_in / 1.1) if supply_vat_in is not None else None
    margin = (
        supply_vat_ex - cost_vat_ex - (r.get("배송비") or 0)
        if supply_vat_ex is not None and cost_vat_ex is not None
        else None
    )
    margin_rate = (margin / supply_vat_ex) if margin is not None and supply_vat_ex else None
    return {
        "event_price": event_price,
        "price_missing": price_missing,
        "할인율": discount_rate,
        "할인금액": discount_amt,
        "원가(+VAT)": cost_vat_in,
        "원가(-VAT)": cost_vat_ex,
        "공급가(+VAT)": supply_vat_in,
        "공급가(-VAT)": supply_vat_ex,
        "공헌이익": margin,
        "공헌이익률": margin_rate,
    }


MONEY_COLUMNS = {
    "정상가(판매가)", "최저가", "상시할인가", "7일 행사가(폐쇄몰)", "3일 행사가",
    "공동구매가", "적용행사가", "할인금액", "원가(+VAT)", "원가(-VAT)",
    "수수료(%)", "배송비", "공급가(+VAT)", "공급가(-VAT)", "공헌이익",
}


def build_styled_excel(df, missing_flags, recent_flags=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "행사제안서"

    headers = list(df.columns)
    ws.append(headers)

    header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    missing_fill = PatternFill(start_color="FFF3B0", end_color="FFF3B0", fill_type="solid")
    data_font = Font(size=10)
    red_font = Font(size=10, bold=True, color="D33333")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    red_thick = Side(style="medium", color="D33333")
    red_border = Border(left=red_thick, right=red_thick, top=red_thick, bottom=red_thick)

    recent_flags = recent_flags or []
    sku_col_idx = headers.index("품번") + 1 if "품번" in headers else None

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_vals in df.itertuples(index=False):
        ws.append(list(row_vals))

    for r in range(2, ws.max_row + 1):
        is_missing = missing_flags[r - 2] if (r - 2) < len(missing_flags) else False
        is_recent = recent_flags[r - 2] if (r - 2) < len(recent_flags) else False
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.font = data_font
            header = headers[c - 1]
            cell.alignment = Alignment(
                horizontal="right" if header in MONEY_COLUMNS else "left", vertical="center"
            )
            if header in MONEY_COLUMNS and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
            if is_missing:
                cell.fill = missing_fill
        if is_recent and sku_col_idx:
            sku_cell = ws.cell(row=r, column=sku_col_idx)
            sku_cell.border = red_border
            sku_cell.font = red_font

    for idx, h in enumerate(headers, start=1):
        col_vals = [str(v) for v in df[h].tolist()]
        max_len = max([len(str(h))] + [len(v) for v in col_vals]) if col_vals else len(str(h))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 4, 10), 32)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def pick_default_type(price):
    """가격이 채워진 항목 중 우선순위대로 기본 적용유형을 고른다."""
    if not price:
        return "day7"
    for k in ["day7", "day3", "최저가", "상시가", "공동구매가"]:
        if price.get(k) is not None:
            return k
    return "day7"


def iround(v):
    """숫자면 소수점 이하를 버리고 정수로, 아니면 그대로 반환."""
    if isinstance(v, (int, float)):
        return int(round(v))
    return v


# ---------------------------------------------------------------
# 기준 파일 저장/로드
# ---------------------------------------------------------------
def save_uploaded_files(uploaded_files):
    """업로드된 파일을 확장자에 따라 data/cost 또는 data/price 폴더에 저장(같은 이름이면 덮어씀)."""
    for f in uploaded_files:
        ext = f.name.lower().rsplit(".", 1)[-1]
        target_dir = COST_DIR if ext == "xls" else PRICE_DIR
        with open(os.path.join(target_dir, f.name), "wb") as out:
            out.write(f.getvalue())


def load_saved_maps():
    """저장된 폴더의 모든 파일을 읽어 cost_map/price_map과 상태 목록을 만든다."""
    cost_map, price_map, status = {}, {}, []
    for fn in sorted(os.listdir(COST_DIR)):
        path = os.path.join(COST_DIR, fn)
        try:
            with open(path, "rb") as fh:
                rows = read_rows(fh.read(), "xls")
            m = parse_cost_sheet(rows)
            cost_map.update(m)
            status.append((fn, "원가 파일", len(m)))
        except Exception as e:
            status.append((fn, f"⚠️ 읽기 실패: {e}", 0))
    for fn in sorted(os.listdir(PRICE_DIR)):
        path = os.path.join(PRICE_DIR, fn)
        try:
            with open(path, "rb") as fh:
                rows = read_rows(fh.read(), "xlsx")
            m = parse_price_sheet(rows)
            price_map.update(m)
            status.append((fn, "가격표 파일", len(m)))
        except Exception as e:
            status.append((fn, f"⚠️ 읽기 실패: {e}", 0))
    return cost_map, price_map, status


def load_saved_channels():
    """data/channel 폴더의 파일들을 읽어 {판매처명: {오클릭품번: 어드민상품코드}} 딕셔너리를 만든다.
    판매처명은 파일명(확장자 제외)."""
    channel_maps, status = {}, []
    for fn in sorted(os.listdir(CHANNEL_DIR)):
        path = os.path.join(CHANNEL_DIR, fn)
        ext = fn.lower().rsplit(".", 1)[-1]
        channel_name = os.path.splitext(fn)[0]
        try:
            with open(path, "rb") as fh:
                rows = read_rows(fh.read(), ext)
            m = parse_channel_sheet(rows)
            channel_maps[channel_name] = m
            status.append((fn, f"판매처 매핑({channel_name})", len(m)))
        except Exception as e:
            status.append((fn, f"⚠️ 읽기 실패: {e}", 0))
    return channel_maps, status


def load_recent_skus(channel, days=14):
    """최근 N일 이내에 해당 판매처로 다운로드 이력이 있는 품번 집합을 반환."""
    if not channel or not os.path.exists(LOG_PATH):
        return set()
    try:
        log_df = pd.read_csv(LOG_PATH, dtype=str)
    except Exception:
        return set()
    if log_df.empty:
        return set()
    log_df["date"] = pd.to_datetime(log_df["date"], errors="coerce")
    cutoff = pd.Timestamp(date.today()) - pd.Timedelta(days=days)
    recent = log_df[(log_df["channel"] == channel) & (log_df["date"] >= cutoff)]
    return set(recent["sku"].astype(str))


def append_download_log(channel, skus):
    """다운로드한 판매처/품번/날짜를 기록에 남긴다."""
    if not channel or not skus:
        return
    today_str = date.today().isoformat()
    new_rows = pd.DataFrame({"date": [today_str] * len(skus), "channel": [channel] * len(skus), "sku": skus})
    if os.path.exists(LOG_PATH):
        try:
            existing = pd.read_csv(LOG_PATH, dtype=str)
            combined = pd.concat([existing, new_rows], ignore_index=True)
        except Exception:
            combined = new_rows
    else:
        combined = new_rows
    combined.to_csv(LOG_PATH, index=False)


def build_rows_from_skus(skus, cost_map, price_map, prev_rows):
    """품번 리스트를 기준으로 결과 행을 만든다. 기존 행이 있으면 수수료/배송비/적용유형 등 편집값을 유지."""
    prev = {r["품번"]: r for r in prev_rows}
    rows = []
    for sku in skus:
        cost = cost_map.get(sku)
        price = price_map.get(sku)
        p = prev.get(sku, {})
        rows.append(
            {
                "품번": sku,
                "상품코드": sku,
                "품명": cost["품명"] if cost else "",
                "정상가": price["정상가"] if price else None,
                "최저가": price["최저가"] if price else None,
                "상시가": price["상시가"] if price else None,
                "day7": price["day7"] if price else None,
                "day3": price["day3"] if price else None,
                "공동구매가": price["공동구매가"] if price else None,
                "rocket": price["rocket"] if price else "-",
                "arrival": price["arrival"] if price else "-",
                "matched": bool(cost or price),
                "적용타입": p.get("적용타입", pick_default_type(price)),
                "기타금액": p.get("기타금액", None),
                "수수료": p.get("수수료", 0.0),
                "배송비": p.get("배송비", 0.0),
                "선택": p.get("선택", True),
            }
        )
    return rows


def search_by_name(query, cost_map):
    """공백으로 구분된 키워드가 순서 상관없이 모두 포함된 품명을 검색."""
    tokens = [t for t in re.split(r"\s+", query.strip()) if t]
    if not tokens:
        return []
    matches = []
    for code, info in cost_map.items():
        name = info.get("품명", "")
        if all(tok in name for tok in tokens):
            matches.append((code, name))
    return matches


def show_no_match_popup(query):
    """일치하는 품명이 없을 때 확인이 필요하다는 팝업(모달)을 띄운다."""
    if hasattr(st, "dialog"):
        @st.dialog("확인이 필요합니다")
        def _dialog():
            st.write(f"'{query}'와(과) 일치하는 품명을 찾을 수 없습니다.")
            st.write("검색어를 다시 확인해주세요 (예: 띄어쓰기, 오탈자 등).")
            if st.button("닫기"):
                st.rerun()
        _dialog()
    else:
        st.warning(f"'{query}'와(과) 일치하는 품명을 찾을 수 없습니다. 검색어를 다시 확인해주세요.")


# ---------------------------------------------------------------
# UI
# ---------------------------------------------------------------
st.title("행사제안서 자동생성기")
st.caption("기준 파일 업로드 → 품번 입력 → 수수료/배송비 설정 → 엑셀 다운로드")

col1, col2 = st.columns(2)

with col1:
    st.subheader("1. 기준 파일 업로드")
    st.caption(
        ".xls = 원가 파일(오클릭)  |  .xlsx = 가격표 파일. "
        "한 번 업로드하면 서버에 저장되어 다음부터는 다시 올릴 필요가 없습니다. "
        "교체하려면 같은 이름의 새 파일을 다시 업로드하세요."
    )
    uploaded = st.file_uploader(
        "파일을 드래그하거나 클릭하여 업로드",
        type=["xls", "xlsx"],
        accept_multiple_files=True,
        key="uploader",
    )
    if uploaded:
        save_uploaded_files(uploaded)
        st.success(f"{len(uploaded)}개 파일이 저장되었습니다.")

    cost_files_on_disk = sorted(os.listdir(COST_DIR))
    price_files_on_disk = sorted(os.listdir(PRICE_DIR))
    if cost_files_on_disk or price_files_on_disk:
        st.write("**저장된 기준 파일**")
        for fn in cost_files_on_disk:
            c1, c2 = st.columns([5, 1])
            c1.write(f"📄 {fn} (원가)")
            if c2.button("삭제", key=f"del_cost_{fn}"):
                os.remove(os.path.join(COST_DIR, fn))
                st.rerun()
        for fn in price_files_on_disk:
            c1, c2 = st.columns([5, 1])
            c1.write(f"📄 {fn} (가격표)")
            if c2.button("삭제", key=f"del_price_{fn}"):
                os.remove(os.path.join(PRICE_DIR, fn))
                st.rerun()
    else:
        st.info("저장된 기준 파일이 없습니다. 파일을 업로드해주세요.")

cost_map, price_map, file_status = load_saved_maps()
channel_maps, channel_status = load_saved_channels()
file_status = file_status + channel_status

with col1:
    with st.expander("파일 인식 상태 보기"):
        for name, kind, cnt in file_status:
            st.write(f"- **{name}** — {kind} ({cnt}건 인식)")

st.subheader("1-2. 판매처(어드민 상품코드) 매핑 파일 업로드")
st.caption(
    "파일명이 그대로 판매처명이 됩니다 (예: '스마트스토어.xls' 업로드 → 판매처 드롭다운에 '스마트스토어' 추가). "
    "오클릭 품번 ↔ 어드민 상품코드가 들어있는 파일을 올려주세요. 이 파일도 서버에 저장되어 계속 유지됩니다."
)
channel_uploaded = st.file_uploader(
    "판매처 매핑 파일을 드래그하거나 클릭하여 업로드 (파일명 = 판매처명)",
    type=["xls", "xlsx"],
    accept_multiple_files=True,
    key="channel_uploader",
)
if channel_uploaded:
    for f in channel_uploaded:
        with open(os.path.join(CHANNEL_DIR, f.name), "wb") as out:
            out.write(f.getvalue())
    st.success(f"{len(channel_uploaded)}개 판매처 매핑 파일이 저장되었습니다.")
    channel_maps, channel_status = load_saved_channels()

channel_files_on_disk = sorted(os.listdir(CHANNEL_DIR))
if channel_files_on_disk:
    cc1, cc2 = st.columns(2)
    for i, fn in enumerate(channel_files_on_disk):
        target = cc1 if i % 2 == 0 else cc2
        c1, c2 = target.columns([4, 1])
        c1.write(f"🏬 {os.path.splitext(fn)[0]}")
        if c2.button("삭제", key=f"del_channel_{fn}"):
            os.remove(os.path.join(CHANNEL_DIR, fn))
            st.rerun()

channel_options = ["(선택 안 함)"] + sorted(channel_maps.keys())
selected_channel = st.selectbox("판매처 선택", channel_options, key="channel_select")
if selected_channel == "(선택 안 함)":
    selected_channel = None

if "rows" not in st.session_state:
    st.session_state.rows = []
if "name_search_results" not in st.session_state:
    st.session_state.name_search_results = []

with col2:
    st.subheader("2. 품번 / 품명 입력")
    tab_code, tab_name = st.tabs(["품번으로 입력", "품명으로 찾기"])

    with tab_code:
        sku_text = st.text_area(
            "품번을 입력하세요. 엔터(줄바꿈) 또는 콤마로 여러 개 구분",
            height=110,
            placeholder="예) 102649, 102650\n102761",
        )
        lookup = st.button("조회하기", type="primary", key="lookup_by_code")
        if lookup:
            seen, skus = set(), []
            for tok in re.split(r"[\n,，]+", sku_text):
                code = normalize_code(tok)
                if code and code not in seen:
                    seen.add(code)
                    skus.append(code)
            st.session_state.rows = build_rows_from_skus(skus, cost_map, price_map, st.session_state.rows)

    with tab_name:
        name_query = st.text_input(
            "품명 키워드를 입력하세요 (띄어쓰기로 구분, 순서 상관없이 검색됩니다)",
            placeholder="예: 크리스탈 램프",
        )
        search_clicked = st.button("검색", key="search_by_name")
        if search_clicked:
            if not name_query.strip():
                st.warning("검색어를 입력해주세요.")
            else:
                matches = search_by_name(name_query, cost_map)
                if not matches:
                    show_no_match_popup(name_query)
                    st.session_state.name_search_results = []
                else:
                    st.session_state.name_search_results = matches

        if st.session_state.name_search_results:
            st.write(f"검색 결과 {len(st.session_state.name_search_results)}건 — 추가할 품목을 선택하세요.")
            options = [f"{code} | {name}" for code, name in st.session_state.name_search_results]
            selected = st.multiselect("품번 선택", options, key="name_search_select")
            if st.button("선택한 품번 추가", key="add_selected_skus"):
                codes = [s.split(" | ")[0] for s in selected]
                if not codes:
                    st.warning("추가할 품목을 먼저 선택해주세요.")
                else:
                    existing = [r["품번"] for r in st.session_state.rows]
                    merged = existing + [c for c in codes if c not in existing]
                    st.session_state.rows = build_rows_from_skus(merged, cost_map, price_map, st.session_state.rows)
                    st.session_state.name_search_results = []
                    st.rerun()

st.subheader("3. 수수료 · 배송비 일괄 적용")
st.caption("결과표의 '적용' 체크박스가 켜진 행에만 일괄 적용됩니다. 개별 행은 결과표에서 직접 수정할 수 있습니다.")
b1, b2, b3 = st.columns([1, 1, 2])
with b1:
    bulk_fee = st.number_input("수수료 일괄(%)", value=0.0, step=0.5)
with b2:
    bulk_ship = st.number_input("배송비 일괄(원)", value=0.0, step=100.0)
with b3:
    st.write("")
    st.write("")
    if st.button("체크된 행에 적용"):
        for r in st.session_state.rows:
            if r.get("선택", True):
                r["수수료"] = bulk_fee
                r["배송비"] = bulk_ship

st.subheader("4. 결과표")

if not st.session_state.rows:
    st.info("기준 파일을 업로드하고 품번을 조회하면 결과가 표시됩니다.")
else:
    unmatched = [r["품번"] for r in st.session_state.rows if not r["matched"]]
    if unmatched:
        st.warning("기준 파일에서 매칭되지 않은 품번: " + ", ".join(unmatched))

    edit_df = pd.DataFrame(st.session_state.rows)

    st.caption(
        "※ 적용유형에서 금액이 비어있는 항목을 고르면 아래 결과표에 노란색으로 표시됩니다 "
        "(선택 자체는 가능하지만 금액이 없어 계산에서 제외됩니다). "
        "'적용' 체크박스는 수수료·배송비 일괄 적용 대상 여부입니다."
    )

    edited = st.data_editor(
        edit_df,
        column_order=[
            "선택", "품번", "상품코드", "품명", "정상가", "최저가", "상시가",
            "day7", "day3", "공동구매가", "적용타입", "기타금액", "수수료", "배송비",
            "rocket", "arrival",
        ],
        column_config={
            "선택": st.column_config.CheckboxColumn("적용", default=True),
            "품번": "품번",
            "상품코드": "상품코드",
            "품명": "품명",
            "정상가": st.column_config.NumberColumn("정상가(판매가)"),
            "최저가": "최저가",
            "상시가": "상시할인가",
            "day7": "7일 행사가(폐쇄몰)",
            "day3": "3일 행사가",
            "공동구매가": "공동구매가",
            "적용타입": st.column_config.SelectboxColumn("적용유형(선택)", options=EVENT_KEYS, required=True),
            "기타금액": st.column_config.NumberColumn("기타 금액(적용유형=기타일 때)"),
            "수수료": st.column_config.NumberColumn("수수료(%)"),
            "배송비": st.column_config.NumberColumn("배송비"),
            "rocket": st.column_config.TextColumn("로켓배송", disabled=True),
            "arrival": st.column_config.TextColumn("도착배송", disabled=True),
        },
        disabled=["품번", "상품코드", "품명", "정상가", "최저가", "상시가", "day7", "day3", "공동구매가", "rocket", "arrival"],
        hide_index=True,
        use_container_width=True,
        key="editor",
    )
    st.session_state.rows = edited.to_dict("records")

    recent_skus = load_recent_skus(selected_channel, days=14) if selected_channel else set()

    display_rows = []
    missing_flags = []
    recent_flags = []
    for r in st.session_state.rows:
        cost_raw = cost_map.get(r["품번"], {}).get("원가")
        c = calc_row(r, cost_raw)
        missing_flags.append(c["price_missing"])
        recent_flags.append(r["품번"] in recent_skus)
        admin_code = "-"
        if selected_channel:
            admin_code = channel_maps.get(selected_channel, {}).get(r["품번"], "미등록")
        display_rows.append(
            {
                "품번": r["품번"],
                "어드민 상품코드": admin_code,
                "품명": r["품명"],
                "정상가(판매가)": iround(r["정상가"]),
                "최저가": iround(r["최저가"]),
                "상시할인가": iround(r["상시가"]),
                "7일 행사가(폐쇄몰)": iround(r["day7"]),
                "3일 행사가": iround(r["day3"]),
                "공동구매가": iround(r["공동구매가"]),
                "적용행사가": iround(c["event_price"]) if c["event_price"] is not None else "확인필요",
                "할인율": f"{c['할인율']*100:.1f}%" if c["할인율"] is not None else "-",
                "할인금액": iround(c["할인금액"]),
                "원가(+VAT)": iround(c["원가(+VAT)"]),
                "원가(-VAT)": iround(c["원가(-VAT)"]),
                "수수료(%)": iround(r["수수료"]),
                "배송비": iround(r["배송비"]),
                "공급가(+VAT)": iround(c["공급가(+VAT)"]),
                "공급가(-VAT)": iround(c["공급가(-VAT)"]),
                "공헌이익": iround(c["공헌이익"]),
                "공헌이익률": f"{c['공헌이익률']*100:.1f}%" if c["공헌이익률"] is not None else "-",
                "로켓배송": r["rocket"],
                "도착배송": r["arrival"],
            }
        )
    if not selected_channel:
        for d in display_rows:
            del d["어드민 상품코드"]

    result_df = pd.DataFrame(display_rows)

    def highlight_missing(row):
        idx = row.name
        if idx < len(missing_flags) and missing_flags[idx]:
            return ["background-color: #fff3b0"] * len(row)
        return [""] * len(row)

    def highlight_recent(col):
        return ["border: 2px solid #d33; font-weight: 700;" if recent_flags[i] else "" for i in range(len(col))]

    if selected_channel and any(recent_flags):
        st.warning(
            "🔴 빨간 테두리로 표시된 품번은 최근 14일 이내에 같은 판매처('"
            + selected_channel + "')로 이미 엑셀 다운로드한 이력이 있습니다."
        )

    styled = result_df.style.apply(highlight_missing, axis=1)
    if any(recent_flags):
        styled = styled.apply(highlight_recent, subset=["품번"])

    st.dataframe(styled, use_container_width=True, hide_index=True)

    buf = build_styled_excel(result_df, missing_flags, recent_flags)
    channel_label = selected_channel if selected_channel else "전체"
    downloaded = st.download_button(
        "엑셀로 다운로드",
        data=buf.getvalue(),
        file_name=f"행사제안서_{channel_label}_{date.today().isoformat()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    if downloaded and selected_channel:
        append_download_log(selected_channel, [r["품번"] for r in st.session_state.rows])
